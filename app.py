print("Hello World")

# Variable

full_name = "John Smith"
age = 20
is_new = True
bill = 525.52

# Input Taking From User

name = input('What is your name? ')
birth_year = int(input('Birth year: '))
weight = float(input("Your Weight In Pound: "))
convert_kilo = 0.45 * weight
print("Your weight is", convert_kilo, "kg")

# String

course = '''
Hello Akg,

Here is our first email to you.

Thank you,
The support team
'''

print(course)

course = 'Python for Beginners'
print(course[0])
print(course[-1])
hello = "Hello"
print(hello[1:-1])

first_name = 'John'
last_name = 'Smith'
message = first_name + ' [' + last_name + '] is a coder'
msg = f'{first_name} [{last_name}] is a coder'
print(message)
print(msg)

course = 'Python for Beginners'
print(course.upper())
print(course.lower())
print(len(course))
print(course.find('P'))
print(course.replace('P', 'J'))
print('P' in course)

# Arithmetic Operations

print(10 / 3)
print(10 // 3)
print(10 % 3)
print(10 ** 3)

import math

print(round(2.9))
print(abs(-5))
print(math.ceil(2.5))

# If Else Condition

price = 1000000
has_good_credit = False

if has_good_credit:
    down_payment = price * .1

else:
    down_payment = price * 0.2

print(f'Down Payment: ${down_payment}')

name = input("Enter your name: ")

if len(name) < 3:
    print('Name must be at least 3 characters')
elif len(name) > 50:
    print('Name can be a maximum of 50 characters')
else:
    print('Name looks good!')

# While Loop

i = 1
while i <= 5:
    print('*' * i)
    i += 1
print('Done')

# For Loop

for item in 'Python':
    print(item)
for item in range(2, 10, 2):
    print(item)

# List

prices = [10, 20, 30]
total_price = 0
for price in prices:
    total_price += price
print(total_price)

# 2D List

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9],
]
print(matrix[0][0])
for row in matrix:
    for item in row:
        print(item)

numbers = [1, 278, 3, 4, 5, 5, 6, 7]
numbers.append(20)
numbers.insert(2, 5)
print(numbers)
numbers.remove(5)
print(numbers)
numbers.pop()
print(numbers)
print(numbers.index(6))
print(50 in numbers)
print(numbers.count(5))
numbers.sort()
numbers.reverse()
print(numbers)
new_numbers = numbers.copy()
new_numbers.append(88)
print(new_numbers)
numbers.clear()
print(numbers)

# Tuples

numbers = (1, 2, 3)
print(numbers[0])

# Unpacking
coordinates = (1, 2, 3)
x, y, z = coordinates
print(x)

# Dictionaries
customers = {
    "name": "John Smith",
    "age": 20,
    "is_verified": True
}

print(customers["name"])
print(customers.get("birth_date", "Jan 1 1980"))
print(customers)


# Function
def greet_user(f_name):
    print(f'Hi {f_name}')


greet_user("John")

# Exception

try:
    age = int(input("Age: "))
    print(age)
except Exception as exp:
    print(exp)


# Class

class Point:

    def __init__(self, name):
        self.name = name

    def move(self):
        print(f"move {self.name}")

    def draw(self):
        print(f"draw {self.name}")


point1 = Point("John")
point1.draw()
point1.move()


# Inheritance

class Animal:

    def walk(self):
        print("Walk")


class Dog(Animal):
    pass


class Cat(Animal):
    pass


dog = Dog()
dog.walk()

# Modules
# https://docs.python.org/3/py-modindex.html

import utils
from utils import kg_to_lbs

print(kg_to_lbs(75))
print(utils.kg_to_lbs(70))

# Package
from package import module
import package.module

package.module.calling()
module.calling()

# Path
from pathlib import Path

# Absolute Path
# Relative Path

path = Path("package")
print(path.exists())
path_2 = Path("module_2")
path_2.mkdir()
path_2.rmdir()

path = Path()
for file in path.glob('*'):
    print(file)

# Working with OpenPyXl

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
print(sheet.cell(1, 1).value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions_updated.xlsx')
